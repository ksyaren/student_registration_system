from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd 
from openpyxl import Workbook
import pathlib 


background = "#1B1E2B"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path("student_data.xlsx")
if file.exists():
    pass
else: 
    workbook = Workbook()  # Yeni bir çalışma kitabı oluştur
    sheet = workbook.active  
    sheet["A1"] = "Registration No"
    sheet["B1"] = "Name"
    sheet["C1"] = "surname"
    sheet["D1"] = "Gender"
    sheet["E1"] = "Date of Birth"
    sheet["F1"] = "Registration Date"
    sheet["G1"] = "Email"
    sheet["H1"] = "Phone Number"
    sheet["I1"] = "University"
    sheet["J1"] = "Student No"
    sheet["K1"] = "Class"
    sheet["L1"] = "Skill"

    workbook.save("student_data.xlsx")

#exit
def Exit():
    root.destroy()

def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file", filetype=(("JPG File","*.jpg"), ("PNG File","*.png"),
                                                                               ("ALL Files","*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image= photo2)
    lbl.image = photo2

##### registration #############
# design automatic registration no entry system 

def registration_no():
    file = openpyxl.load_workbook("student_data.xlsx")
    sheet = file.active
    row = sheet.max_row

    max_row_value= sheet.cell(row =row, column = 1).value

    try:
        Registration.set(int(max_row_value) + 1)
    except (ValueError, TypeError):
        Registration.set("1")  # Eğer hücre boşsa veya geçersizse 1 olarak başlat 

###### CLEAR ##############

def clear():
    global img
    Name.set("")
    Surname.set("")
    DOB.set("")
    Email.set("")
    Phone.set("")
    UniName.set("")
    StudentNo.set("")
    Class.set("Select Class")
    Skill.set("")

    registration_no()
    SaveButton.config(state="normal")
    img1 = PhotoImage(file ="images/add-user.png")
    lbl.config(image=img1)
    lbl.image = img1

    img = ""

############### SAVE #####################

def save():
    R1 = Registration.get()
    N1 = Name.get()
    S1= Surname.get()
    try:
        G1= gender
    except :
        messagebox.showerror("error", "Select gender!")

    D2= DOB.get()
    D1 = Date.get()
    E1 = Email.get()
    P1 = Phone.get()
    U1 = UniName.get()
    S2 = StudentNo.get()
    C1 = Class.get()
    S3 = Skill.get()

    if N1 =="" or C1 =="Select Class" or D2 =="" or S1=="" or E1=="" or P1=="" or U1=="" or S2=="":
        messagebox.showerror("error", "Few Data Missing!")
    else:
        file= openpyxl.load_workbook("student_data.xlsx")
        sheet= file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=S1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=E1)
        sheet.cell(column=8, row=sheet.max_row, value=P1) 
        sheet.cell(column=9, row=sheet.max_row, value=U1)
        sheet.cell(column=10, row=sheet.max_row, value=S2)
        sheet.cell(column=11, row=sheet.max_row, value=C1)
        sheet.cell(column=12, row=sheet.max_row, value=S3)
        file.save(r"student_data.xlsx")

        try:
            img.save("Student images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile picture is not available")

        messagebox.showinfo("info","Succesfuly data entered.")

        clear()
        registration_no()

##################### SEARCH #################################
def search():
    
    text = Search.get()

    clear()
    SaveButton.config(state="disabled") # after clicking search button save button will be disabled so no one can click

    file= openpyxl.load_workbook("student_data.xlsx")
    sheet=file.active
     
    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            #print(str(name))
            reg_no_position =str(name)[14:-1]
            reg_number= str(name)[15:-1]

            #print(reg_no_position)
            #print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("invalid","Invalid registration number")

    x1 = sheet.cell(row=int(reg_number),column=1).value
    x2 = sheet.cell(row=int(reg_number),column=2).value
    x3 = sheet.cell(row=int(reg_number),column=3).value
    x4 = sheet.cell(row=int(reg_number),column=4).value
    x5 = sheet.cell(row=int(reg_number),column=5).value
    x6 = sheet.cell(row=int(reg_number),column=6).value
    x7 = sheet.cell(row=int(reg_number),column=7).value
    x8 = sheet.cell(row=int(reg_number),column=8).value
    x9 = sheet.cell(row=int(reg_number),column=9).value
    x10 = sheet.cell(row=int(reg_number),column=10).value
    x11 = sheet.cell(row=int(reg_number),column=11).value
    x12 = sheet.cell(row=int(reg_number),column=12).value

     #print(x1)
    #print(x2)
    #print(x3)
    #print(x4)
    #print(x5)
    #print(x6)
    #print(x7)
    #print(x8)
    #print(x9)
    #print(x10)
    #print(x11) 

    Registration.set(x1)
    Name.set(x2)
    Surname.set(x3)

    if x4 == "Female":
        radio2.select()
    else:
        radio1.select()

    DOB.set(x5)
    Date.set(x6)
    Email.set(x7)
    Phone.set(x8)
    UniName.set(x9)
    StudentNo.set(x10)
    Class.set(x11)
    Skill.set(x12)

    img = (Image.open("Student images/"+ str(x1)+".jpg"))
    resized_images= img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_images)
    lbl.config(image=photo2)
    lbl.image= photo2

def update():
    R1 = Registration.get()
    N1 = Name.get()
    S1= Surname.get()
    selection()
    G1=gender
    D2= DOB.get()
    D1 = Date.get()
    E1 = Email.get()
    P1 = Phone.get()
    U1 = UniName.get()
    S2 = StudentNo.get()
    C1 = Class.get()
    S3 = Skill.get()

    file =openpyxl.load_workbook("student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name=row[0]
            print(str(name))
            reg_no_position =str(name)[14:-1]
            reg_number = str(name)[15:-1]
            print(reg_number)

    #sheet.cell(column =1, row=int(reg_number),value=R1) no one can update registration no
    sheet.cell(column =2, row=int(reg_number),value=N1)
    sheet.cell(column =3, row=int(reg_number),value=S1)
    sheet.cell(column =4, row=int(reg_number),value=G1)
    sheet.cell(column =5, row=int(reg_number),value=D2)
    sheet.cell(column =6, row=int(reg_number),value=D1)
    sheet.cell(column =7, row=int(reg_number),value=E1)
    sheet.cell(column =8, row=int(reg_number),value=P1)
    sheet.cell(column =9, row=int(reg_number),value=U1)
    sheet.cell(column =10, row=int(reg_number),value=S2)
    sheet.cell(column =11, row=int(reg_number),value=C1)
    sheet.cell(column =12, row=int(reg_number),value=S3)

    file.save(r"student_data.xlsx")
    
    try:
        img.save("Student images/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("update","Updated successfully")
    


#gender
def selection():
    global gender
    value=radio.get()
    if value == 1:
        gender = "Male"
    else :
        gender = "Female"

# top frames
Label(
   root,
    text="Student Registration System",
    bg="#292D3E",  # Daha koyu bir arka plan
    fg="white",
    font=("Arial", 20, "bold"),  # Daha büyük ve modern font
    pady=15,  # Padding yüksekliği
).pack(side=TOP, fill=X)



# search box
Search = StringVar()
Entry(root, textvariable= Search, width=10, bd=2, bg="#f8f8f6", font="arial 19 ").place(x =910, y =80)
imageicon3 =PhotoImage(file ="images/search.png")
Srch= Button(root, text="Search student ",bg="#f8f8f6", fg= "#1B1E2B",compound=LEFT ,image = imageicon3, 
             width =130, height= 28, command=search)
Srch.place(x= 1070, y =80)

imageicon4 =PhotoImage(file ="images/reload.png")
update_button = Button(root, image = imageicon4, width =55, height= 30, bg="#292D3E", command= update)
update_button.place(x= 100, y =18)


# registration  and date
Label(
   root,
    text="Registration No: ",
    bg="#1B1E2B",  # Daha koyu bir arka plan
    fg="#f8f8f6",
    font=("Arial", 13 ),  # Daha büyük ve modern font
).place(x= 30, y=130)

Label(
   root,
    text="Date: ",
    bg="#1B1E2B",  # Daha koyu bir arka plan
    fg="#f8f8f6",
    font=("Arial", 13 ),  # Daha büyük ve modern font
).place(x= 480, y=130)

Registration= IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable= Registration, width=15, font=("Arial", 10), bg="#f8f8f6")
reg_entry.place(x= 160, y=130)

registration_no()


today = date.today()
d1=today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable= Date, width=15, font=("Arial 10"),bg="#f8f8f6")
date_entry.place(x=550, y=130)

Date.set(d1)

#students_details
obj = LabelFrame(root, text="Student's Details", 
                 font=("Arial 16"), bd=2, 
                 bg="#292D3E",
                 fg="#f8f8f6",
                 width=850,
                height=230, relief= GROOVE)
obj.place(x=30, y=200)

Label( obj,text="Name: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 30, y=50)
Label( obj,text="Surname: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x=400, y=50)
Label( obj,text="Gender: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 30, y=150)


Label( obj,text="Date of Birth: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 30, y=100)
Label( obj,text="Email: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x=400, y=100)
Label( obj,text="Phone Number: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 400, y=150)

Name = StringVar()
name_entry= Entry(obj, textvariable= Name, width= 20, font=("Arial", 10),bg="#f8f8f6")
name_entry.place(x=100, y=50)

Surname = StringVar()
surname_entry= Entry(obj, textvariable= Surname, width= 20, font=("Arial", 10),bg="#f8f8f6")
surname_entry.place(x=490, y=50)  # Surname alanı doğru yere taşındı

radio= IntVar()
radio1 = Radiobutton(obj, text="Male", variable=radio, bg="#f8f8f6",value=1, command=selection) 
radio1.place(x=120, y=150)

radio2 = Radiobutton(obj, text="Female", bg="#f8f8f6", variable=radio, value=2, command=selection)
radio2.place(x=180, y=150)

DOB = StringVar()
dob_entry= Entry(obj, textvariable= DOB ,width= 20,bg="#f8f8f6", font=("Arial", 10))
dob_entry.place(x=140, y=100)  # DOB alanı doğru yere taşındı

Email = StringVar()
email_entry= Entry(obj, textvariable= Email, bg="#f8f8f6",width= 20, font=("Arial", 10))
email_entry.place(x=460, y=100)

Phone = StringVar()
phone_entry= Entry(obj, textvariable= Phone, bg="#f8f8f6",width= 20, font=("Arial", 10))
phone_entry.place(x=530, y=150)

#school_details
obj2 = LabelFrame(root, text="University's Details", 
                 font=("Arial 16"), bd=2, 
                 bg="#292D3E",
                 fg="#f8f8f6",
                 width=850,
                height=200, relief= GROOVE)
obj2.place(x=30, y=450)

Label( obj2,text="University Name: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 30, y=40)
Label( obj2,text="Student No: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 30, y=100)

Label( obj2,text="Class: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 400, y=40)
Label( obj2,text="Skill: ",bg="#292D3E", fg="#f8f8f6",font=("Arial", 13 ),  ).place(x= 400, y=100)

UniName = StringVar()
uniName_entry= Entry(obj2, textvariable= UniName, width= 20, font=("Arial", 10),bg="#f8f8f6" )
uniName_entry.place(x=170, y=40)

StudentNo = StringVar()
student_no_entry= Entry(obj2, textvariable= StudentNo, width= 20, font=("Arial", 10),bg="#f8f8f6")
student_no_entry.place(x=130, y=100)

Class = Combobox(obj2, values=["1","2","3","4"], font=("Arial",10), width =17, state = "r"  )
Class.place(x=460, y=40)
Class.set("Select Class")

Skill = StringVar()
skill_entry= Entry(obj2, textvariable= Skill, width= 20, font=("Arial", 10),bg="#f8f8f6")
skill_entry.place(x=450, y=100)


#image

f =Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE) 
f.place(x=980, y=150)

img = PhotoImage(file="images/add-user.png")
lbl = Label(f, bg="black",image= img)
lbl.place(x=0, y=0)


#button

Button(root, text="Upload", width=19, height=2, font="Arial 12 bold",bg="#88CE65", command=showimage).place(x=980, y=370)
SaveButton=Button(root, text="Save", width=19, height=2, font="Arial 12 bold", bg="#17ACE8", command = save)
SaveButton.place(x=980, y=450)
Button(root, text="Reset", width=19, height=2, font="Arial 12 bold",bg="#FFBB17", command=clear).place(x=980, y=530)
Button(root, text="Exit", width=19, height=2, font="Arial 12 bold",bg="#F44336", command=Exit).place(x=980, y=610)

root.mainloop()

